import random
from openpyxl import Workbook
from openpyxl.styles import Alignment
import datetime
import time

def enter():
    return print('\n')

enter()
print("인원 셔플과 출석부 제작을 해주는 프로그램입니다.")

people = ["강려령", "강보경", "고요한", "김윤성",
          "김준호", "나정연", "송창현", "안우리",
          "유찬호", "최지수", "유태영", "이지윤",
          "이윤형", "유새람", "이찬하", "박정현",
          "윤형규", "민동준", "정예은", "문상호",
          "이혜인", "임다은", "임지민", "임형준",
          "최형규", "이호동", "홍창남", "김재훈",
          "함형우", "조승윤", "지혜정", "전수진"]

print("현재 포함되어 있는 사람들입니다.")
enter()
people.sort()
print(people)
prev_total_people = len(people)

# 제외되는 사람 입력하고 셔플돌리기
enter()
print("띄어쓰기 단위로 끊어서 제외할 사람을 입력해주세요.(없으면 enter)")
out_of_group = input()

# 띄어쓰기 단위로 배열 나누기
out_of_group = out_of_group.split()
# 제외된 사람 기존 리스트에서 지우기
for i in out_of_group:
    if i in people:
        people.remove(i)

after_total_people = len(people)

if prev_total_people == after_total_people:
    print("제외된 사람이 없습니다. 전원참석 ^^")
else:
    print("입력한 사람을 제외한 결과입니다.")

enter()
print(people)

enter()
# 묶는 인원 지정
print("몇 명씩 묶을까요?")
grouping = input()
grouping = int(grouping)

# 인원 섞기
print("쉐킷쉐킷~ 섞는 중입니다.\n")
time.sleep(3)
people = random.sample(people, len(people))

# 그룹별 출력
for i in range(len(people)):
    print(people[i] + ', ', end="")
    if i != 0 and i % grouping == grouping - 1:
        enter()
# 출석부 export
enter()
print('랜덤 배정이 완료되었습니다.')
print('해당 내용을 엑셀로 저장하시겠습니까?(y/n)')
answer = input()

if answer == 'y' or answer == "Y":
    wb = Workbook()

    ws = wb.active
    ws.column_dimensions[chr(grouping + 65)].width = 30.0

    ws['A1'] = "참여자"
    ws['A1'].alignment = Alignment(horizontal="center")
    checktime = chr(grouping + 65) + '3'
    ws[checktime] = datetime.datetime.now()

    # 칼럼별로 이름 writing
    index_list = []
    i = 0
    while i < after_total_people / grouping:
        index_list.append(chr(i + 65))
        i = i + 1

    j = 0
    for i in range(len(people)):
        index = index_list[i % grouping] + str(j + 2)
        ws[index] = people[i]
        ws[index].alignment = Alignment(horizontal="center")
        if i % grouping == grouping - 1:
            j = j + 1

    # 날짜 형식 0801 처럼 만들기
    today = datetime.date.today()
    if today.month < 10:
        month = str('0') + str(today.month)
    else:
        month = str(today.month)

    if today.day < 10:
        day = str('0') + str(today.day)
    else:
        day = str(today.day)

    date = month + day

    wb.save('출석부' + date + '.xlsx')
    print("저장되었습니다.")
